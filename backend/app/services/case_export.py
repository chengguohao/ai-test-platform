"""用例树 ↔ XMind / Excel 导出与回读（共用用例树 JSON 结构，见方案 §七）。"""
from __future__ import annotations

import json
import uuid
import zipfile
from pathlib import Path

# ---------------- 用例树规范工具 ----------------

PREFIX_KEYS = {
    "前置条件": "precondition", "前置": "precondition",
    "测试数据": "data", "数据": "data",
    "优先级": "priority",
    "关联接口": "api", "接口": "api",
    "备注": "remark",
}


def _split_prefix(title: str) -> tuple[str, str]:
    """按「前缀: 内容」拆分；无前缀返回 ("", title)。"""
    for i, ch in enumerate(title):
        if ch == ":" or ch == "：":
            key = title[:i].strip()
            if key:
                return key, title[i + 1:].strip()
    return "", title


def _strip_seq(s: str) -> str:
    """去掉步骤/预期合并分支回读时的序号前缀（如 "1) xxx" → "xxx"）。"""
    import re
    return re.sub(r"^\s*\d+\)\s*", "", s).strip()


def _case_node(case: dict) -> list[dict]:
    """把一条用例拆成 5 个详情子分支：前置条件/步骤/预期结果/优先级/关联接口。

    步骤、预期结果为合并单节点：主干标题"步骤："/"预期结果："，下挂 1 个子节点，
    子节点标题为多行文本 "1) a\n2) b\n3) c"；空字段整组省略。
    """
    nodes: list[dict] = []
    pre = (case.get("precondition") or "").strip()
    if pre:
        nodes.append(_topic(f"前置条件：{pre}"))

    steps = [s for s in (case.get("steps") or []) if s and s.strip()]
    if steps:
        step_text = "\n".join(f"{i}) {s}" for i, s in enumerate(steps, 1))
        nodes.append(_topic("步骤：", [_topic(step_text)]))

    expects = [e for e in (case.get("expects") or []) if e and e.strip()]
    if expects:
        expect_text = "\n".join(f"{i}) {e}" for i, e in enumerate(expects, 1))
        nodes.append(_topic("预期结果：", [_topic(expect_text)]))

    pri = (case.get("priority") or "").strip()
    if pri:
        nodes.append(_topic(f"优先级：{pri}"))

    api = (case.get("api") or "").strip()
    if api:
        nodes.append(_topic(f"关联接口：{api}"))

    return nodes


# ---------------- XMind（现代 content.json 格式，纯标准库） ----------------

def _estimate_width(text: str) -> int:
    """按内容长度估算 XMind 节点宽度（像素），实现「长度根据内容扩充」。

    中文字符/全角标点按 14px、ASCII 按 8px 累计，取最长行；加左右内边距 32px，
    最小 96px（XMind 默认最小宽度）。
    """
    max_w = 0
    for line in str(text).split("\n"):
        w = sum(14 if ord(ch) > 0x2E7F else 8 for ch in line)
        max_w = max(max_w, w)
    return max(96, max_w + 32)


def _topic(title: str, children: list | None = None) -> dict:
    t = {
        "id": "topic-" + uuid.uuid4().hex,
        "class": "topic",
        "title": title,
        "width": _estimate_width(title),  # 显式宽度=内容长度，避免 XMind 固定默认宽度
    }
    if children:
        t["children"] = {"attached": children}
    return t


def export_xmind(tree: dict, path: Path, root_title: str | None = None) -> Path:
    """用例树 → .xmind（content.json 树 + metadata.json + manifest.json）。

    结构：根主题=本次需求名称（root_title）→ 分组/模块名分支 → 用例名称分支 →
    单个详情节点（前置条件/步骤/预期结果 各一行，合并展示便于阅读）。
    """
    root = _topic(root_title or tree.get("title", "测试用例"))
    root_children = []
    for g in tree.get("groups", []):
        cases = [_topic(f"{c.get('id', '')} {c.get('title', '')}", _case_node(c))
                 for c in g.get("cases", [])]
        root_children.append(_topic(g.get("name", "分组"), cases))
    root["children"] = {"attached": root_children}

    sheet = {
        "id": "sheet-" + uuid.uuid4().hex,
        "class": "sheet",
        "title": "Sheet 1",
        "rootTopic": root,
        "topicPositioning": "fixed",
        "theme": {"name": "Classic"},
        "extensions": [],
    }
    content = json.dumps([sheet], ensure_ascii=False)
    metadata = json.dumps({"creator": {"name": "ai-test-platform", "version": "1.0.0"}}, ensure_ascii=False)
    manifest = json.dumps({"file-entries": {"content.json": {}, "metadata.json": {}}}, ensure_ascii=False)

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("content.json", content.encode("utf-8"))
        zf.writestr("metadata.json", metadata.encode("utf-8"))
        zf.writestr("manifest.json", manifest.encode("utf-8"))
    return path


def _topic_children(topic: dict) -> list:
    return (topic.get("children") or {}).get("attached") or []


def parse_xmind(path: Path) -> dict:
    """.xmind → 用例树 JSON（容错：新用例自动补编号，无前缀子节点归步骤）。"""
    with zipfile.ZipFile(path) as zf:
        content = json.loads(zf.read("content.json").decode("utf-8"))
    sheet = content[0]
    root = sheet.get("rootTopic", {})
    module = _auto_module(root.get("title", ""))
    tree = {
        "module": module, "title": root.get("title", "手工测试用例"),
        "groups": [], "version": 1,
    }
    used_ids: set[str] = set()
    counter = 0
    # module 优先从用例 ID 前缀提取（TC-XXX-…），根主题为中文需求名时标题提取不到
    module_from_ids = ""

    for group_topic in _topic_children(root):
        group = {"name": group_topic.get("title", "分组"), "cases": []}
        for case_topic in _topic_children(group_topic):
            title = case_topic.get("title", "").strip()
            cid, ctitle = _case_id_title(title, module, used_ids, counter)
            if not module_from_ids and cid.startswith("TC-"):
                parts = cid.split("-")
                if len(parts) >= 2 and parts[1]:
                    module_from_ids = parts[1].lower()
            used_ids.add(cid)
            counter += 1
            case = {"id": cid, "title": ctitle, "precondition": "", "data": "",
                    "steps": [], "expects": [], "priority": "", "api": "", "remark": ""}
            # 子节点可能：
            # 用户图示格式（合并单节点）：5 个主干 + 步骤/预期各下挂 1 个子节点
            #   子节点标题为多行文本 "1) a\n2) b\n3) c"
            # 旧版多节点：前置/步骤i/预期i 各成独立子节点
            for child in _topic_children(case_topic):
                child_title = child.get("title", "")
                key, val = _split_prefix(child_title)
                # 1) 步骤/预期主干：标题仅"步骤："/"预期结果："，下挂 1 个多行子节点
                if (key in ("步骤", "预期结果")
                        or child_title.strip() in ("步骤：", "预期结果：")):
                    target = "steps" if key == "步骤" else "expects"
                    subs = _topic_children(child)
                    if subs:
                        # 单子节点多行文本：按行拆出去序号前缀
                        for line in [l for l in subs[0].get("title", "").split("\n") if l.strip()]:
                            case[target].append(_strip_seq(line))
                    else:
                        # 兼容旧版：主干直接是多行文本
                        for line in [l for l in child_title.split("\n") if l.strip()]:
                            case[target].append(_strip_seq(line))
                    continue
                # 2) 旧版多节点：主干标题为"步骤1：xxx"/"预期结果1：xxx"
                if key.startswith("步骤") or key.startswith("step"):
                    case["steps"].append(_strip_seq(val))
                    continue
                if key.startswith("预期") or key.startswith("expect"):
                    case["expects"].append(_strip_seq(val))
                    continue
                # 3) 前置/优先级/关联接口：val 可能是"无"，还原为空
                if key in PREFIX_KEYS and PREFIX_KEYS[key] != "":
                    case[PREFIX_KEYS[key]] = "" if val == "无" else val
                else:
                    # 兜底：无前缀整行归入步骤
                    case["steps"].append(_strip_seq(child_title))
            # 步骤/预期数量不对称时补齐空串，保证可读
            n = max(len(case["steps"]), len(case["expects"]))
            case["steps"] += [""] * (n - len(case["steps"]))
            case["expects"] += [""] * (n - len(case["expects"]))
            group["cases"].append(case)
        tree["groups"].append(group)
    # 根主题（中文需求名）提取不到英文模块名时，回退用例 ID 前缀（TC-OA_NOTICE-…）
    if module == "module" and module_from_ids:
        tree["module"] = module_from_ids
    return tree


def _auto_module(title: str) -> str:
    """从标题粗提取模块短名（找不到用 module）。"""
    import re
    m = re.search(r"([A-Za-z][A-Za-z0-9_]{0,20})", title)
    return m.group(1).lower() if m else "module"


def _case_id_title(raw: str, module: str, used: set, counter: int) -> tuple[str, str]:
    """解析用例标题：TC- 前缀则用原标题，否则自动补编号。"""
    raw = raw.strip()
    parts = raw.split(" ", 1)
    if parts and parts[0].startswith("TC-"):
        cid, ctitle = parts[0], (parts[1].strip() if len(parts) > 1 else parts[0])
    else:
        cid = f"TC-{module.upper()}-{counter + 1:02d}"
        ctitle = raw or cid
    # 冲突时自动追加后缀
    while cid in used:
        cid = f"{cid}-{counter + 1}"
    return cid, ctitle


# ---------------- Excel（openpyxl） ----------------

EXCEL_HEADERS = ["用例ID", "用例名称", "模块", "优先级", "前置条件", "测试数据",
                 "步骤", "预期结果", "关联接口", "备注"]


def export_excel(tree: dict, path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(EXCEL_HEADERS)
    header_fill = PatternFill("solid", fgColor="4B3FE3")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    module = tree.get("module", "")
    for g in tree.get("groups", []):
        for c in g.get("cases", []):
            ws.append([c.get("id", ""), c.get("title", ""), g.get("name", module),
                       c.get("priority", ""), c.get("precondition", ""), c.get("data", ""),
                       "\n".join(c.get("steps", [])), "\n".join(c.get("expects", [])),
                       c.get("api", ""), c.get("remark", "")])
    for col in ("A", "B", "E", "F"):
        ws.column_dimensions[col].width = 18
    for col in ("G", "H", "J"):
        ws.column_dimensions[col].width = 30
    wb.save(path)
    return path


def parse_excel(path: Path) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h or "").strip() for h in rows[0]]
    tree = {"module": "", "title": f"{Path(path).stem} 手工测试用例", "groups": [], "version": 1}
    group_map: dict[str, list] = {}
    for row in rows[1:]:
        if not any(row):
            continue
        rec = {header[i]: (str(row[i] or "").strip()) for i in range(min(len(header), len(row)))}
        gname = rec.get("模块") or "默认分组"
        group_map.setdefault(gname, []).append({
            "id": rec.get("用例ID") or "",
            "title": rec.get("用例名称") or "",
            "precondition": rec.get("前置条件") or "",
            "data": rec.get("测试数据") or "",
            "steps": [s for s in (rec.get("步骤") or "").splitlines() if s],
            "expects": [s for s in (rec.get("预期结果") or "").splitlines() if s],
            "priority": rec.get("优先级") or "",
            "api": rec.get("关联接口") or "",
            "remark": rec.get("备注") or "",
        })
    tree["groups"] = [{"name": k, "cases": v} for k, v in group_map.items()]
    return tree
